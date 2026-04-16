"""Populate the Ryder Quote Template with extracted items."""

import shutil
import logging
from copy import copy
from datetime import date, datetime
from pathlib import Path
import openpyxl
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
import config
from pipeline.extractor import ExtractionResult

logger = logging.getLogger(__name__)


def _copy_row_style(ws, src_row: int, tgt_row: int) -> None:
    """Copy cell formatting from src_row to tgt_row across all 6 columns."""
    for col in range(1, 7):
        src = ws.cell(row=src_row, column=col)
        tgt = ws.cell(row=tgt_row, column=col)
        if src.has_style:
            tgt.font = copy(src.font)
            tgt.border = copy(src.border)
            tgt.fill = copy(src.fill)
            tgt.alignment = copy(src.alignment)
            tgt.number_format = src.number_format


def _update_sheet1(wb, vehicle_title: str) -> None:
    """Insert logo and set dynamic processing date in Sheet1."""
    try:
        ws1 = wb["Sheet1"]

        # --- Dynamic processing date in D8 (format: 14-Apr-26) ---
        ws1["D8"] = date.today().strftime("%d-%b-%y")

    except Exception as e:
        logger.warning(f"Sheet1 update failed: {e}")


def write_populated_workbook(result: ExtractionResult, output_path: str) -> str:
    """
    Copy master template, populate SPO Quote sheet with extracted data,
    update Sheet1 with logo and dynamic date,
    preserve rows 31-57, clear pricing values.
    Returns the output path.
    """
    template = str(config.TEMPLATE_PATH)
    shutil.copy2(template, output_path)

    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb[config.SPO_SHEET]

    # --- Sheet1: logo + dynamic date ---
    _update_sheet1(wb, result.vehicle_title or "")

    # --- SPO Quote: Row 2 — empty spacer (keep A2="C", clear C2/D2) ---
    ws.cell(row=2, column=config.CODE_COL).value = None
    ws.cell(row=2, column=config.DESC_COL).value = None

    # --- SPO Quote: Row 3 — vehicle title ---
    ws.cell(row=3, column=config.ACTION_COL).value = "C"
    ws.cell(row=3, column=config.DESC_COL).value = result.vehicle_title or ""

    # --- SPO Quote: Row 33 — upfit item line, clear for manual entry ---
    for col in range(1, 7):
        ws.cell(row=33, column=col).value = None

    # --- SPO Quote: Rows 4+ — extracted items ---
    items = result.items
    for i, item in enumerate(items):
        row = config.DATA_START_ROW + i

        # If we're about to overwrite the preserved section, insert a row first
        if row >= config.PRESERVE_FROM_ROW:
            ws.insert_rows(config.PRESERVE_FROM_ROW)

        ws.cell(row=row, column=config.ACTION_COL).value = "C"
        ws.cell(row=row, column=config.CODE_COL).value = item.code
        ws.cell(row=row, column=config.DESC_COL).value = item.description

        # Copy style from the row above
        if row > config.DATA_START_ROW:
            _copy_row_style(ws, row - 1, row)

    # --- Clear pricing values (keep labels in col D, blank col E) ---
    # Search by label text — row positions may have shifted if rows were inserted
    for row_idx in range(config.PRESERVE_FROM_ROW, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=config.DESC_COL).value
        if label in ("Total", "Upfit Total", "Grand Total"):
            ws.cell(row=row_idx, column=config.PRICE_COL).value = None

    wb.save(output_path)
    logger.info(f"Workbook saved: {output_path}")
    return output_path


def build_output_path(pdf_path: str) -> str:
    """Derive output Excel path from input PDF filename, with timestamp to avoid conflicts."""
    stem = Path(pdf_path).stem
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return str(config.OUTPUT_DIR / f"{stem} - POPULATED - {ts}.xlsx")
